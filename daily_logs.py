import streamlit as st
import pandas as pd
import requests
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import plotly.express as px

# ---------------------- Constants ----------------------
EXCEL_FILE = "generated_reports.xlsx"
PANEL_FILE = "lt_panel_checklist.xlsx"
LT_PANEL_FILE = "lt_panel_log.xlsx"
CHILLER_EXCEL_FILE = "chiller_readings.xlsx"
HF_API_KEY = os.getenv("HF_API_KEY")  # Or use st.secrets
HF_API_URL = "https://api-inference.huggingface.co/models/mistralai/Mistral-7B-Instruct-v0.1"

# ---------------------- Utility Functions ----------------------
def apply_excel_styling(file_path):
    """Apply consistent styling to Excel files"""
    try:
        wb = load_workbook(file_path)
        thin_border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"),
            top=Side(style="thin"), 
            bottom=Side(style="thin")
        )
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            
            # Apply borders and alignment
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            
            # Bold headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = max(len(str(cell.value)) for cell in col if cell.value)
                adjusted_width = min(max_length + 2, 50)  # Cap width at 50
                ws.column_dimensions[col[0].column_letter].width = adjusted_width
        
        wb.save(file_path)
        return True
    except Exception as e:
        st.error(f"Error applying styling: {e}")
        return False

# ---------------------- Maintenance Report Generator ----------------------
def generate_report(unit, machine, technician_name, issue):
    """Generate maintenance report using AI model"""
    prompt = f"""
As an expert electrical maintenance engineer, create a concise one-line report:
Unit: {unit}
Machine: {machine}
Technician: {technician_name}
Issue: {issue}
- Strictly one line only
- Include problem and solution
- Professional tone
"""
    headers = {"Authorization": f"Bearer {HF_API_KEY}", "Content-Type": "application/json"}
    payload = {
        "inputs": prompt,
        "parameters": {"max_new_tokens": 100, "temperature": 0.7, "return_full_text": False}
    }

    try:
        response = requests.post(HF_API_URL, headers=headers, json=payload, timeout=30)
        if response.status_code == 200:
            return response.json()[0]["generated_text"].strip()
        return f"Resolved by {technician_name}: {issue}"
    except Exception as e:
        st.warning(f"AI service unavailable: {e}")
        return f"Resolved by {technician_name}: {issue}"

def save_to_excel(unit, machine, technician_name, issue, report):
    """Save report to Excel with proper formatting"""
    new_data = pd.DataFrame([{
        "Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Unit": unit, 
        "Machine": machine,
        "Technician": technician_name, 
        "Issue": issue,
        "Report": report
    }])
    
    if os.path.exists(EXCEL_FILE):
        try:
            existing_data = pd.read_excel(EXCEL_FILE)
            updated_data = pd.concat([existing_data, new_data], ignore_index=True)
        except:
            updated_data = new_data
    else:
        updated_data = new_data

    updated_data.to_excel(EXCEL_FILE, index=False)
    apply_excel_styling(EXCEL_FILE)

def maintenance_report_ui():
    """UI for maintenance report generation"""
    st.title("📝 AI-Powered Maintenance Report Generator")
    
    with st.form("report_form"):
        col1, col2 = st.columns(2)
        with col1:
            unit = st.text_input("Unit:", placeholder="e.g., Unit 5", key="unit")
            machine = st.text_input("Machine:", placeholder="e.g., Compressor A1", key="machine")
        with col2:
            technician = st.text_input("Technician:", placeholder="e.g., John Doe", key="tech")
            issue = st.text_input("Issue:", placeholder="e.g., High temperature", key="issue")
        
        submitted = st.form_submit_button("Generate Report")
        if submitted:
            if all([unit, machine, technician, issue]):
                with st.spinner("Generating report..."):
                    report = generate_report(unit, machine, technician, issue)
                    save_to_excel(unit, machine, technician, issue, report)
                    st.success("✅ Report saved!")
                    st.subheader("Generated Report:")
                    st.info(report)
            else:
                st.warning("Please fill all fields")

    st.divider()
    st.subheader("📁 Report History")
    
    if os.path.exists(EXCEL_FILE):
        df = pd.read_excel(EXCEL_FILE)
        st.dataframe(df, use_container_width=True, hide_index=True)
        with open(EXCEL_FILE, "rb") as f:
            st.download_button("📥 Download Reports", f, file_name="maintenance_reports.xlsx")
    else:
        st.info("No reports generated yet")

# ---------------------- LT Panel Readings Entry ----------------------
def lt_panel_input(panel_name):
    """Input fields for LT panel readings"""
    st.subheader(panel_name)
    cols = st.columns(4)
    with cols[0]:
        volt = st.number_input("Volt", min_value=0.0, key=f"{panel_name}_volt")
    with cols[1]:
        amp = st.number_input("Amp", min_value=0.0, key=f"{panel_name}_amp")
    with cols[2]:
        pf = st.number_input("PF", min_value=0.0, max_value=1.0, value=0.98, step=0.01, key=f"{panel_name}_pf")
    with cols[3]:
        temp = st.number_input("Temp (°C)", min_value=0.0, key=f"{panel_name}_temp")
    return volt, amp, pf, temp

def save_lt_panel_data(date, shift, time, reader, readings):
    """Save LT panel data to Excel"""
    if os.path.exists(LT_PANEL_FILE):
        wb = load_workbook(LT_PANEL_FILE)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    for panel_name, values in readings.items():
        row_data = {
            "Date": date.strftime("%Y-%m-%d"),
            "Shift": shift,
            "Time": time.strftime("%H:%M"),
            "Technician": reader,
            "Volt": values[0],
            "Amp": values[1],
            "PF": values[2],
            "Temp": values[3]
        }

        df = pd.DataFrame([row_data])
        
        if panel_name in wb.sheetnames:
            ws = wb[panel_name]
            for r_idx, row in df.iterrows():
                ws.append(row.tolist())
        else:
            ws = wb.create_sheet(panel_name)
            ws.append(list(df.columns))
            for r_idx, row in df.iterrows():
                ws.append(row.tolist())

    wb.save(LT_PANEL_FILE)
    apply_excel_styling(LT_PANEL_FILE)

def show_lt_panel_trend():
    """Visualize LT panel trends"""
    st.title("📈 LT Panel Trends")
    
    if not os.path.exists(LT_PANEL_FILE):
        st.warning("No data available")
        return

    try:
        wb = load_workbook(LT_PANEL_FILE, data_only=True)
        panels = wb.sheetnames
        selected_panels = st.multiselect("Select Panels", panels, default=panels[:2])
        parameter = st.selectbox("Parameter", ["Volt", "Amp", "PF", "Temp"])
        
        if not selected_panels:
            return
            
        combined_df = pd.DataFrame()
        
        for panel in selected_panels:
            ws = wb[panel]
            data = ws.values
            cols = next(data)
            df = pd.DataFrame(data, columns=cols)
            
            # Convert to numeric and handle errors
            df[parameter] = pd.to_numeric(df[parameter], errors='coerce')
            df['Timestamp'] = pd.to_datetime(df['Date'] + ' ' + df['Time'], errors='coerce')
            
            df = df.dropna(subset=['Timestamp', parameter])
            df['Panel'] = panel
            combined_df = pd.concat([combined_df, df])
            
        if combined_df.empty:
            st.warning("No valid data to display")
            return
            
        fig = px.line(
            combined_df, 
            x='Timestamp', 
            y=parameter, 
            color='Panel',
            markers=True,
            title=f"{parameter} Trend"
        )
        st.plotly_chart(fig, use_container_width=True)
        
    except Exception as e:
        st.error(f"Error: {e}")

def lt_panel_ui():
    """UI for LT panel data entry"""
    st.title("📊 LT Panel Monitoring")
    
    if os.path.exists(LT_PANEL_FILE):
        with open(LT_PANEL_FILE, "rb") as f:
            st.download_button("📥 Download Data", f, file_name="lt_panel_data.xlsx")
    
    if st.button("Show Trends"):
        show_lt_panel_trend()

    with st.form("lt_form"):
        st.subheader("General Information")
        cols = st.columns(3)
        with cols[0]:
            date = st.date_input("Date", value=datetime.today())
        with cols[1]:
            shift = st.selectbox("Shift", ["A", "B", "C"])
        with cols[2]:
            time = st.time_input("Time", value=datetime.now().time())
        reader = st.text_input("Technician", placeholder="Your name")
        
        st.divider()
        st.subheader("Panel Readings")
        
        panels = ["LT Panel 1", "LT Panel 2", "LT Panel 3", 
                 "LT Panel 4", "Tapline", "Looms Panel"]
        readings = {}
        
        for panel in panels:
            readings[panel] = lt_panel_input(panel)
        
        submitted = st.form_submit_button("Save Data")
        if submitted:
            save_lt_panel_data(date, shift, time, reader, readings)
            st.success("✅ Data saved!")

# ---------------------- Compressor Readings ----------------------
def compressor_excel_logger():
    """UI for compressor data logging"""
    st.title("💨 Compressor Monitoring")
    
    EXCEL_PATH = "compressor_log.xlsx"
    
    if os.path.exists(EXCEL_PATH):
        with open(EXCEL_PATH, "rb") as f:
            st.download_button("📥 Download Logs", f, file_name="compressor_logs.xlsx")

    compressors = {
        "Unit 1 (37 KW)": {},
        "Unit 2 (37 KW)": {},
        "Unit 3 (55 KW)": {},
        "Unit 4 (55 KW)": {},
        "Unit 5 (22 KW)": {}
    }

    with st.form("compressor_form"):
        st.subheader("Log Details")
        cols = st.columns(2)
        with cols[0]:
            log_date = st.date_input("Date", value=datetime.today())
        with cols[1]:
            shift = st.selectbox("Shift", ["A", "B", "C"])
        
        st.divider()
        st.subheader("Compressor Readings")
        
        for unit in compressors.keys():
            st.markdown(f"**{unit}**")
            cols = st.columns(3)
            with cols[0]:
                amps = st.number_input("Amps", min_value=0.0, key=f"{unit}_amps")
            with cols[1]:
                temp = st.number_input("Temp (°C)", min_value=0.0, key=f"{unit}_temp")
            with cols[2]:
                pressure = st.number_input("Pressure (bar)", min_value=0.0, key=f"{unit}_pres")
            
            compressors[unit] = {"Amps": amps, "TempC": temp, "PressBar": pressure}
        
        submitted = st.form_submit_button("Save Log")
        if submitted:
            try:
                if os.path.exists(EXCEL_PATH):
                    wb = load_workbook(EXCEL_PATH)
                else:
                    wb = Workbook()
                    wb.remove(wb.active)

                for unit, data in compressors.items():
                    sheet_name = unit.split(" ")[0] + unit.split(" ")[1]
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                    else:
                        ws = wb.create_sheet(title=sheet_name)
                        ws.append(["Date", "Shift", "Amps", "Temp (°C)", "Pressure (bar)"])
                    
                    ws.append([
                        log_date.strftime("%Y-%m-%d"),
                        shift,
                        data["Amps"],
                        data["TempC"],
                        data["PressBar"]
                    ])
                
                wb.save(EXCEL_PATH)
                apply_excel_styling(EXCEL_PATH)
                st.success("✅ Log saved!")
                
            except Exception as e:
                st.error(f"Error: {e}")

# ---------------------- Chiller Readings ----------------------
CHILLER_NAMES = {
    "CHILLER 1 (UNIT 1)": "Chiller1",
    "CHILLER 2 (BACKUP)": "Chiller2",
    "CHILLER 3 (UNIT 2)": "Chiller3",
    "CHILLER 4 (UNIT 3)": "Chiller4",
    "CHILLER 5 (BACKUP)": "Chiller5",
    "CHILLER 6 (UNIT 3)": "Chiller6",
    "CHILLER 7 (UNIT 4)": "Chiller7",
    "CHILLER 8 (BACKUP)": "Chiller8",
    "CHILLER 9 (UNIT 6)": "Chiller9",
    "CHILLER 10 (UNIT 8)": "Chiller10",
    "CHILLER 11 (UNIT 8)": "Chiller11", 
}

def save_chiller_data(shift, time, chiller_readings):
    """Save chiller data to Excel"""
    time_str = time.strftime("%I:%M %p")
    
    if os.path.exists(CHILLER_EXCEL_FILE):
        wb = load_workbook(CHILLER_EXCEL_FILE)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    for chiller_name, readings in chiller_readings.items():
        sheet_name = CHILLER_NAMES[chiller_name]
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["SHIFT", "TIME", "AMP", "COOLING TEMP", "PRESSURE", "OIL LEVEL"])
        
        ws.append([shift, time_str] + readings)
    
    wb.save(CHILLER_EXCEL_FILE)
    apply_excel_styling(CHILLER_EXCEL_FILE)

def show_chiller_trend():
    """Visualize chiller trends"""
    st.title("📊 Chiller Performance Trends")
    
    if not os.path.exists(CHILLER_EXCEL_FILE):
        st.warning("No data available")
        return
        
    try:
        wb = load_workbook(CHILLER_EXCEL_FILE, data_only=True)
        chillers = wb.sheetnames
        selected = st.multiselect("Select Chillers", chillers, default=chillers[:3])
        parameter = st.selectbox("Parameter", ["AMP", "COOLING TEMP", "PRESSURE", "OIL LEVEL"])
        
        if not selected:
            return
            
        combined = pd.DataFrame()
        
        for chiller in selected:
            ws = wb[chiller]
            data = ws.values
            cols = next(data)
            df = pd.DataFrame(data, columns=cols)
            
            # Convert to numeric and handle errors
            df[parameter] = pd.to_numeric(df[parameter], errors='coerce')
            df['TIME'] = pd.to_datetime(df['TIME'], format='%I:%M %p', errors='coerce')
            
            df = df.dropna(subset=['TIME', parameter])
            df['Chiller'] = chiller
            combined = pd.concat([combined, df])
            
        if combined.empty:
            st.warning("No valid data to display")
            return
            
        fig = px.line(
            combined, 
            x='TIME', 
            y=parameter, 
            color='Chiller',
            markers=True,
            title=f"{parameter} Trend"
        )
        st.plotly_chart(fig, use_container_width=True)
        
    except Exception as e:
        st.error(f"Error: {e}")

def chiller_ui():
    """UI for chiller monitoring"""
    st.title("❄️ Chiller Monitoring")
    
    if os.path.exists(CHILLER_EXCEL_FILE):
        with open(CHILLER_EXCEL_FILE, "rb") as f:
            st.download_button("📥 Download Data", f, file_name="chiller_data.xlsx")
    
    if st.button("Show Trends"):
        show_chiller_trend()

    with st.form("chiller_form"):
        cols = st.columns(2)
        with cols[0]:
            shift = st.selectbox("Shift", ["A", "B", "C"])
        with cols[1]:
            time = st.time_input("Time", value=datetime.now().time())
        
        st.divider()
        st.subheader("Chiller Readings")
        
        readings = {}
        for name in CHILLER_NAMES:
            st.markdown(f"**{name}**")
            cols = st.columns(4)
            with cols[0]:
                amp = st.text_input("Amp", key=f"{name}_amp")
            with cols[1]:
                temp = st.text_input("Cooling Temp", key=f"{name}_temp")
            with cols[2]:
                pressure = st.text_input("Pressure", key=f"{name}_pressure")
            with cols[3]:
                oil = st.text_input("Oil Level", key=f"{name}_oil")
            
            readings[name] = [amp, temp, pressure, oil]
        
        submitted = st.form_submit_button("Save Readings")
        if submitted:
            save_chiller_data(shift, time, readings)
            st.success("✅ Readings saved!")

# ---------------------- Main App ----------------------
def main():
    st.set_page_config(
        page_title="Maintenance Dashboard", 
        page_icon="🔧", 
        layout="wide"
    )
    
    st.sidebar.title("🔧 Maintenance Dashboard")
    tool = st.sidebar.radio("Select Tool", [
        "Report Generator", 
        "LT Panel", 
        "Compressors",
        "Chillers"
    ])
    
    st.sidebar.divider()
    st.sidebar.info("Maintenance Management System v1.0")
    
    if tool == "Report Generator":
        maintenance_report_ui()
    elif tool == "LT Panel":
        lt_panel_ui()
    elif tool == "Compressors":
        compressor_excel_logger()
    elif tool == "Chillers":
        chiller_ui()

if __name__ == "__main__":
    main()
